#!/usr/bin/env python3
"""
Instamart Keyword Dashboard
============================
Ingests Granular CSV reports, appends to a master log, and outputs
an Excel dashboard with per-keyword stats and target bid recommendations.

Usage:
  python dashboard.py --input report.csv
  python dashboard.py --input report.csv --target-roas 4.0
  python dashboard.py --input report.csv --epoch-start 2026-08-10
  python dashboard.py --input report.csv --min-bid 2.5 --target-roas 4.0

Arguments:
  --input         Path to new Granular CSV report (required)
  --master        Path to master log CSV (default: master_log.csv)
  --output        Output Excel file (default: keyword_dashboard.xlsx)
  --target-roas   Target ROAS for bid calculation (default: 3.0)
  --epoch-start   Only use data from this date onwards (YYYY-MM-DD).
                  Use when you change a listing. Overrides the 30-day window.
  --min-bid       Minimum bid on Swiggy Instamart (default: None).
                  Keywords with target bid below this will be flagged.
  --window        Rolling window in days (default: 30). Ignored if --epoch-start is set.
"""

import argparse
import os
import sys
import warnings
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from pathlib import Path

warnings.filterwarnings("ignore")

# ── Constants ────────────────────────────────────────────────────────────────

HEADER_ROWS = 6
MIN_DAYS    = 7      # minimum observations to compute reliable stats
MIN_CLICKS  = 50     # minimum total clicks for binomial SE to be meaningful

COLS_TO_SUM = [
    "TOTAL_IMPRESSIONS",
    "TOTAL_CLICKS",
    "TOTAL_BUDGET_BURNT",
    "TOTAL_A2C",
    "TOTAL_CONVERSIONS",
    "TOTAL_GMV",
]

DEDUP_KEY = ["METRICS_DATE", "PRODUCT_NAME", "KEYWORD", "CITY"]


# ── 1. Load new CSV ───────────────────────────────────────────────────────────

def load_csv(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, skiprows=HEADER_ROWS, dtype=str)
    df.columns = df.columns.str.strip()

    # Keep only the columns we need
    keep = DEDUP_KEY + COLS_TO_SUM + ["CAMPAIGN_NAME", "CAMPAIGN_ID"]
    keep = [c for c in keep if c in df.columns]
    df = df[keep].copy()

    # Coerce numeric columns
    for col in COLS_TO_SUM:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    df["METRICS_DATE"]  = pd.to_datetime(df["METRICS_DATE"], errors="coerce")
    df["PRODUCT_NAME"]  = df["PRODUCT_NAME"].str.strip()
    df["KEYWORD"]       = df["KEYWORD"].str.strip()

    return df.dropna(subset=["METRICS_DATE", "PRODUCT_NAME", "KEYWORD"])


# ── 2. Master log management ──────────────────────────────────────────────────

def update_master(new_df: pd.DataFrame, master_path: str) -> pd.DataFrame:
    if os.path.exists(master_path):
        master = pd.read_csv(master_path, dtype=str)
        master.columns = master.columns.str.strip()
        for col in COLS_TO_SUM:
            if col in master.columns:
                master[col] = pd.to_numeric(master[col], errors="coerce").fillna(0)
        master["METRICS_DATE"] = pd.to_datetime(master["METRICS_DATE"], errors="coerce")

        # Deduplicate: only append rows not already in master
        existing_keys = set(
            master[DEDUP_KEY].apply(lambda r: tuple(r.values), axis=1)
        )
        new_keys = new_df[DEDUP_KEY].apply(lambda r: tuple(r.values), axis=1)
        new_rows = new_df[~new_keys.isin(existing_keys)]

        if len(new_rows) == 0:
            print("  No new rows — all already in master log.")
        else:
            print(f"  Appended {len(new_rows):,} new rows to master log.")
            master = pd.concat([master, new_rows], ignore_index=True)
    else:
        master = new_df.copy()
        print(f"  Created new master log with {len(master):,} rows.")

    master.to_csv(master_path, index=False)
    return master


# ── 3. Apply epoch / rolling window filter ────────────────────────────────────

def apply_window(master: pd.DataFrame, epoch_start: str | None, window: int) -> pd.DataFrame:
    if epoch_start:
        cutoff = pd.to_datetime(epoch_start)
        label  = f"epoch from {epoch_start}"
    else:
        cutoff = pd.Timestamp.now().normalize() - timedelta(days=window)
        label  = f"last {window} days"

    filtered = master[master["METRICS_DATE"] >= cutoff]
    print(f"  Using data from {label}: {len(filtered):,} rows, "
          f"{filtered['METRICS_DATE'].nunique()} dates.")
    return filtered


# ── 4. Aggregate across cities per date ───────────────────────────────────────

def aggregate_by_date(df: pd.DataFrame) -> pd.DataFrame:
    agg = (
        df.groupby(
            ["METRICS_DATE", "CAMPAIGN_ID", "CAMPAIGN_NAME", "PRODUCT_NAME", "KEYWORD"],
            as_index=False,
        )[COLS_TO_SUM].sum()
    )

    # Daily derived metrics
    agg["CTR"]      = np.where(agg["TOTAL_IMPRESSIONS"] > 0,
                               agg["TOTAL_CLICKS"] / agg["TOTAL_IMPRESSIONS"], np.nan)
    agg["A2C_RATE"] = np.where(agg["TOTAL_CLICKS"] > 0,
                               agg["TOTAL_A2C"] / agg["TOTAL_CLICKS"], np.nan)
    agg["CVR"]      = np.where(agg["TOTAL_CLICKS"] > 0,
                               agg["TOTAL_CONVERSIONS"] / agg["TOTAL_CLICKS"], np.nan)
    agg["eCPC"]     = np.where(agg["TOTAL_CLICKS"] > 0,
                               agg["TOTAL_BUDGET_BURNT"] / agg["TOTAL_CLICKS"], np.nan)
    agg["AOV"]      = np.where(agg["TOTAL_CONVERSIONS"] > 0,
                               agg["TOTAL_GMV"] / agg["TOTAL_CONVERSIONS"], np.nan)
    agg["ROAS"]     = np.where(agg["TOTAL_BUDGET_BURNT"] > 0,
                               agg["TOTAL_GMV"] / agg["TOTAL_BUDGET_BURNT"], np.nan)
    return agg



# ── 5. Build summary (one row per product × keyword) ─────────────────────────

def build_summary(agg: pd.DataFrame, target_roas: float, min_bid: float | None) -> pd.DataFrame:

    # Metrics that use day-based std dev
    day_metrics = {
        "Impressions": "TOTAL_IMPRESSIONS",
        "Clicks":      "TOTAL_CLICKS",
        "Spend":       "TOTAL_BUDGET_BURNT",
        "A2C":         "TOTAL_A2C",
        "Conversions": "TOTAL_CONVERSIONS",
        "GMV":         "TOTAL_GMV",
        "eCPC":        "eCPC",
        "AOV":         "AOV",
        "ROAS":        "ROAS",
    }

    rows = []

    for (campaign_name, product, keyword), grp in agg.groupby(
        ["CAMPAIGN_NAME", "PRODUCT_NAME", "KEYWORD"]
    ):
        grp = grp.sort_values("METRICS_DATE")
        n_days = len(grp)

        # Cumulative totals (for binomial SE and overall rates)
        total_impressions  = grp["TOTAL_IMPRESSIONS"].sum()
        total_clicks       = grp["TOTAL_CLICKS"].sum()
        total_a2c          = grp["TOTAL_A2C"].sum()
        total_conversions  = grp["TOTAL_CONVERSIONS"].sum()
        total_gmv          = grp["TOTAL_GMV"].sum()
        total_spend        = grp["TOTAL_BUDGET_BURNT"].sum()

        row = {
            "Campaign":    campaign_name,
            "Product":     product,
            "Keyword":     keyword,
            "Days":        n_days,
            "Total Clicks": int(total_clicks),
        }

        # ── Day-based stats for absolute / level metrics ──
        for label, col in day_metrics.items():
            series = grp[col].dropna()
            n      = len(series)
            mean   = series.mean() if n > 0 else np.nan
            sd     = series.std(ddof=1) if n > 1 else np.nan
            se     = sd / np.sqrt(n) if n > 1 else np.nan
            row[f"{label} Mean"] = mean
            row[f"{label} SD"]   = sd
            row[f"{label} SE"]   = se

        # ── Binomial SE for proportions ──
        # CTR: p = total_clicks / total_impressions, N = total_impressions
        if total_impressions > 0:
            ctr = total_clicks / total_impressions
            se_ctr = np.sqrt(ctr * (1 - ctr) / total_impressions)
        else:
            ctr, se_ctr = np.nan, np.nan

        # A2C Rate: p = total_a2c / total_clicks, N = total_clicks
        if total_clicks > 0:
            a2c_rate = total_a2c / total_clicks
            se_a2c   = np.sqrt(a2c_rate * (1 - a2c_rate) / total_clicks)
        else:
            a2c_rate, se_a2c = np.nan, np.nan

        # CVR: p = total_conversions / total_clicks, N = total_clicks
        if total_clicks > 0:
            cvr    = total_conversions / total_clicks
            se_cvr = np.sqrt(cvr * (1 - cvr) / total_clicks)
        else:
            cvr, se_cvr = np.nan, np.nan

        row["CTR"]        = ctr
        row["CTR SE"]     = se_ctr
        row["A2C Rate"]   = a2c_rate
        row["A2C Rate SE"]= se_a2c
        row["CVR"]        = cvr
        row["CVR SE"]     = se_cvr

        # Overall ROAS
        row["Overall ROAS"] = total_gmv / total_spend if total_spend > 0 else np.nan

        # ── Target bid calculation ──
        # Target Bid = CVR × AOV / target_ROAS
        # AOV from cumulative totals
        mean_aov = total_gmv / total_conversions if total_conversions > 0 else np.nan
        # SE(AOV) from day-based series
        aov_series = grp["AOV"].dropna()
        se_aov = (aov_series.std(ddof=1) / np.sqrt(len(aov_series))
                  if len(aov_series) > 1 else np.nan)

        if cvr is not np.nan and not np.isnan(cvr) and mean_aov is not np.nan and not np.isnan(mean_aov):
            target_bid = (cvr * mean_aov) / target_roas

            # Error propagation: Var(CVR × AOV) ≈ AOV²·Var(CVR) + CVR²·Var(AOV)
            var_cvr = se_cvr ** 2 if se_cvr and not np.isnan(se_cvr) else 0
            var_aov = se_aov ** 2 if se_aov and not np.isnan(se_aov) else 0
            se_product = np.sqrt(mean_aov**2 * var_cvr + cvr**2 * var_aov)
            se_target_bid = se_product / target_roas

            row["Target Bid"]       = target_bid
            row["Target Bid Low"]   = target_bid - 1.96 * se_target_bid  # 95% CI
            row["Target Bid High"]  = target_bid + 1.96 * se_target_bid
            row["Target Bid SE"]    = se_target_bid
        else:
            row["Target Bid"]       = np.nan
            row["Target Bid Low"]   = np.nan
            row["Target Bid High"]  = np.nan
            row["Target Bid SE"]    = np.nan

        row["Target ROAS"] = target_roas

        # ── Flags ──
        flags = []
        if n_days < MIN_DAYS:
            flags.append(f"Low days ({n_days}<{MIN_DAYS})")
        if total_clicks < MIN_CLICKS:
            flags.append(f"Low clicks ({int(total_clicks)}<{MIN_CLICKS})")
        if (min_bid is not None
                and not np.isnan(row["Target Bid"])
                and row["Target Bid"] < min_bid):
            flags.append(f"Below min bid (₹{min_bid})")
        row["Flags"] = "; ".join(flags) if flags else "OK"

        rows.append(row)

    return pd.DataFrame(rows)


# ── 6. Write Excel ────────────────────────────────────────────────────────────

def write_excel(summary: pd.DataFrame, daily: pd.DataFrame,
                master: pd.DataFrame, output_path: str,
                target_roas: float, min_bid: float | None):

    from openpyxl.styles import (PatternFill, Font, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    from openpyxl import load_workbook

    # Colour palette
    DARK    = "1A1A2E"
    MID     = "16213E"
    ACCENT  = "0F3460"
    GREEN   = "198754"
    RED     = "DC3545"
    YELLOW  = "FFC107"
    WHITE   = "FFFFFF"
    LGREY   = "F8F9FA"

    def header_style(ws, fill_hex=DARK):
        fill = PatternFill("solid", fgColor=fill_hex)
        font = Font(color=WHITE, bold=True, size=10, name="Calibri")
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30

    def auto_width(ws, max_w=28):
        for col in ws.columns:
            length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 2, max_w)

    def thin_border():
        s = Side(style="thin", color="DDDDDD")
        return Border(left=s, right=s, top=s, bottom=s)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # ── Sheet 1: Keyword Summary ──────────────────────────────────────────
        summary.to_excel(writer, sheet_name="Keyword Summary", index=False)
        ws1 = writer.sheets["Keyword Summary"]
        header_style(ws1, DARK)

        # Alternate row shading + borders
        for i, row in enumerate(ws1.iter_rows(min_row=2, max_row=ws1.max_row), start=2):
            fill = PatternFill("solid", fgColor=LGREY) if i % 2 == 0 else PatternFill("solid", fgColor=WHITE)
            for cell in row:
                cell.fill = fill
                cell.border = thin_border()
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Format percentage columns
        pct_cols = ["CTR", "CTR SE", "A2C Rate", "A2C Rate SE", "CVR", "CVR SE"]
        for col_cell in ws1[1]:
            if col_cell.value in pct_cols:
                col_idx = col_cell.column
                for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row,
                                          min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.number_format = "0.00%"

        # Format currency columns
        currency_cols = ["Spend Mean", "GMV Mean", "eCPC Mean", "AOV Mean",
                         "Target Bid", "Target Bid Low", "Target Bid High", "Target Bid SE"]
        for col_cell in ws1[1]:
            if col_cell.value in currency_cols:
                col_idx = col_cell.column
                for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row,
                                          min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.number_format = '₹#,##0.00'

        # Colour-scale on ROAS column
        for col_cell in ws1[1]:
            if col_cell.value == "Overall ROAS":
                col_letter = get_column_letter(col_cell.column)
                ws1.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws1.max_row}",
                    ColorScaleRule(
                        start_type="min", start_color="DC3545",
                        mid_type="num",   mid_value=target_roas, mid_color="FFC107",
                        end_type="max",   end_color="198754",
                    )
                )

        # Colour Flags column
        for col_cell in ws1[1]:
            if col_cell.value == "Flags":
                col_letter = get_column_letter(col_cell.column)
                ws1.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws1.max_row}",
                    CellIsRule(operator="equal", formula=['"OK"'],
                               fill=PatternFill("solid", fgColor="D4EDDA"))
                )
                ws1.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{ws1.max_row}",
                    CellIsRule(operator="notEqual", formula=['"OK"'],
                               fill=PatternFill("solid", fgColor="F8D7DA"))
                )


        auto_width(ws1)
        ws1.freeze_panes = "D2"  # freeze campaign + product + keyword

        # ── Sheet 2: Daily Data ───────────────────────────────────────────────
        daily_out = daily.copy()
        daily_out["METRICS_DATE"] = daily_out["METRICS_DATE"].dt.strftime("%Y-%m-%d")
        daily_out.to_excel(writer, sheet_name="Daily Data", index=False)
        ws2 = writer.sheets["Daily Data"]
        header_style(ws2, MID)
        auto_width(ws2)
        ws2.freeze_panes = "A2"

        # ── Sheet 3: Master Log ───────────────────────────────────────────────
        master_out = master.copy()
        master_out["METRICS_DATE"] = pd.to_datetime(
            master_out["METRICS_DATE"], errors="coerce"
        ).dt.strftime("%Y-%m-%d")
        master_out.to_excel(writer, sheet_name="Master Log", index=False)
        ws3 = writer.sheets["Master Log"]
        header_style(ws3, ACCENT)
        auto_width(ws3)
        ws3.freeze_panes = "A2"

    print(f"  Dashboard saved → {output_path}")


# ── 7. Main ───────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Instamart Keyword Dashboard",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--input",       required=True,        help="New Granular CSV report")
    parser.add_argument("--master",      default="master_log.csv", help="Master log path")
    parser.add_argument("--output",      default="keyword_dashboard.xlsx", help="Output Excel")
    parser.add_argument("--target-roas", type=float, default=3.0, help="Target ROAS (default 3.0)")
    parser.add_argument("--epoch-start", default=None, help="Use data from this date (YYYY-MM-DD)")
    parser.add_argument("--min-bid",     type=float, default=None, help="Swiggy minimum bid amount")
    parser.add_argument("--window",      type=int, default=30, help="Rolling window days (default 30)")
    args = parser.parse_args()

    print("\n── Step 1: Loading CSV ──────────────────────────────────")
    new_df = load_csv(args.input)
    print(f"  Loaded {len(new_df):,} rows | "
          f"Date range: {new_df['METRICS_DATE'].min().date()} → {new_df['METRICS_DATE'].max().date()}")

    print("\n── Step 2: Updating master log ─────────────────────────")
    master = update_master(new_df, args.master)

    print("\n── Step 3: Applying data window ────────────────────────")
    filtered = apply_window(master, args.epoch_start, args.window)

    print("\n── Step 4: Aggregating by date ─────────────────────────")
    daily = aggregate_by_date(filtered)
    print(f"  {len(daily):,} product×keyword×date rows")

    print("\n── Step 5: Building summary ────────────────────────────")
    summary = build_summary(daily, args.target_roas, args.min_bid)
    print(f"  {len(summary):,} product×keyword combinations")

    ok   = (summary["Flags"] == "OK").sum()
    warn = len(summary) - ok
    print(f"  ✓ {ok} OK  ⚠ {warn} flagged")

    print("\n── Step 6: Writing Excel ───────────────────────────────")
    write_excel(summary, daily, master, args.output, args.target_roas, args.min_bid)

    print(f"\n✓ Done.")
    print(f"  Keywords tracked : {len(summary):,}")
    print(f"  Date range       : {filtered['METRICS_DATE'].min().date()} → {filtered['METRICS_DATE'].max().date()}")
    print(f"  Target ROAS      : {args.target_roas}x")
    if args.min_bid:
        print(f"  Minimum bid      : ₹{args.min_bid}")
    if args.epoch_start:
        print(f"  Epoch start      : {args.epoch_start}")


if __name__ == "__main__":
    main()
